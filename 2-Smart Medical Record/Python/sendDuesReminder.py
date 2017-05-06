import openpyxl
import smtplib
import sys

   
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_highest_column()

unpaidMembers = {}

for r in range(2, sheet.get_highest_row()+1):
      payment = sheet.cell(row=r, column=lastCol).value
      if payment != 'paid':
         name = sheet.cell(row=r, column=1).value
         email = sheet.cell(row=r, column=2).value
         unpaidMembers[name] = email

         
#login details     
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(' ashishsardana21@gmail.com ','aufweidersehen')


for name, email in unpaidMembers.items():
     body = "Subject:  Pay your fee for the event..!!"
     print('Sending email to %s...' % email)
     sendmailStatus = smtpObj.sendmail('ashishsardana21@gmail', email, body)
     if sendmailStatus != {}:
           print('There was a problem sending email to %s: %s' % (email,
           sendmailStatus))

smtpObj.quit()
    

    
    
    
   
    
    

       
