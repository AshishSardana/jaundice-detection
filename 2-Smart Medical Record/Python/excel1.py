# Importing libraries
import openpyxl
import smtplib
import sys
from openpyxl.utils import get_column_letter, column_index_from_string

# Initial constants
breakfastCalorie = 95
lunchCalorie = 150
dinnerCalorie = 135
wb = Workbook()
sheet = wb.get_active_sheet()

maxRows = sheet.max_row
name = {}
email = {}
currentP = 0
currentM = ''

# Initializing total for 4 weeks
totalIndex = 11
for i in range(4):
    for index in range(2, maxRows-2, 3):
        sheet[get_column_letter(totalIndex)+str(index+2)] = int(0)
    totalIndex = totalIndex + 1

totalIndex = 11

i = 1
# Storing names
for index in range(2, maxRows-2, 3):
    name[i] = sheet.cell(row = index, column = 1).value
    i = i + 1

i = 1
# Storing emails
for index in range(2, maxRows-2, 3):
    email[i] = sheet.cell(row = index, column = 2).value
    i = i + 1

def rowNo(currentM):
    if currentM == 'b':
            return 2
    elif currentM == 'l':
            return 3
    elif currentM == 'd':
            return 4

def calorie(currentM):
    if currentM == 'b':
            return breakfastCalorie
    elif currentM == 'l':
            return lunchCalorie
    elif currentM == 'd':
            return dinnerCalorie
        
# Writing values for Day-1
def writeToDay(day, currentP, currentM):
    if (currentP==1):
        sheet[get_column_letter(2+day) + str(rowNo(currentM))] = calorie(currentM)
    else:
        sheet[get_column_letter(2+day) + str(rowNo(currentM) + 3*(currentP-1))] = calorie(currentM)

# Total calories
def totalCalories(totalIndex):  
    for index in range(2, maxRows-2, 3):
        sheet['J'+str(index+2)] = '=SUM(\'C\'+str(index):\'E\'+str(index+2))'
        sheet[get_column_letter(totalIndex)+str(index+2)] = sheet['J'+str(index+2)].value        
    totalIndex = totalIndex + 1
